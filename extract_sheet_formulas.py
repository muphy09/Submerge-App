import openpyxl

# Open the Master Excel Pricing workbook
wb = openpyxl.load_workbook('Master Excel Pricing.xlsx', data_only=False)

# Extract formulas from specific sheets
sheets_to_check = ['PLUM', 'ELEC', 'STEEL', 'SHOT', 'EXC']

for sheet_name in sheets_to_check:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*100}")
        print(f"SHEET: {sheet_name}")
        print(f"{'='*100}\n")

        # Read first 50 rows
        for row_idx in range(1, 51):
            has_data = False
            row_data = []

            for col_idx, col in enumerate(['A', 'B', 'C', 'D', 'E', 'F'], start=1):
                cell = ws[f'{col}{row_idx}']
                if cell.value:
                    has_data = True
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        row_data.append(f"[{col}]={cell.value}")
                    else:
                        row_data.append(f"[{col}]:{cell.value}")

            if has_data:
                print(f"Row {row_idx:3d}: " + " | ".join(row_data))

print("\n" + "="*100)
print("EXTRACTION COMPLETE")
print("="*100)
