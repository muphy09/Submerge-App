import openpyxl
from openpyxl.utils import get_column_letter
import json

def extract_inputs_and_costs(file_path):
    """Extract inputs from NEW POOL tab and costs from SUMMARY - NEW tab"""
    wb = openpyxl.load_workbook(file_path, data_only=True)  # data_only=True to get calculated values

    print("=" * 100)
    print(f"EXTRACTING DATA FROM: {file_path}")
    print("=" * 100)

    # Get inputs from NEW POOL tab
    if 'NEW POOL' in wb.sheetnames:
        ws = wb['NEW POOL']
        print("\n### INPUT VALUES FROM NEW POOL TAB ###\n")

        # Define the input fields we want to extract based on Excel analysis
        inputs = {
            'customerName': ws['B3'].value,
            'city': ws['B4'].value,
            'siltFence': ws['B5'].value,
            'perimeter': ws['B7'].value,
            'maxWidth': ws['E7'].value,
            'surfaceArea': ws['B8'].value,
            'maxLength': ws['E8'].value,
            'decking': ws['E9'].value,
            'shallowDepth': ws['B9'].value,
            'endDepth': ws['B10'].value,
            'travel': ws['E10'].value,
            'poolToStreet': ws['E11'].value,
            'totalStepAndBench': ws['B11'].value,
            'tanningShelf': ws['B12'].value,
            'approxGallons': ws['E13'].value,
            'spaLength': ws['B14'].value,
            'spaWidth': ws['B15'].value,
            'spaShape': ws['B16'].value,
            'spaPerimeter': ws['B17'].value,
            'raisedSpa': ws['B18'].value,
            'raisedSpaFacing': ws['B19'].value,
        }

        for key, value in inputs.items():
            print(f"{key}: {value}")

        # Get RBB values
        print("\n### RBB (Raised Bond Beam) VALUES ###\n")
        rbb_data = {
            '6inch': {'lnft': ws['B32'].value, 'facing': ws['C32'].value},
            '12inch': {'lnft': ws['B33'].value, 'facing': ws['C33'].value},
            '18inch': {'lnft': ws['B34'].value, 'facing': ws['C34'].value},
            '24inch': {'lnft': ws['B35'].value, 'facing': ws['C35'].value},
            '30inch': {'lnft': ws['B36'].value, 'facing': ws['C36'].value},
            '36inch': {'lnft': ws['B37'].value, 'facing': ws['C37'].value},
            'totalSqft': ws['B38'].value,
        }
        for key, value in rbb_data.items():
            print(f"{key}: {value}")

        # Get more inputs
        print("\n### ADDITIONAL INPUTS ###\n")
        additional = {
            'columns': ws['B40'].value,
            'columnWidth': ws['B41'].value,
            'columnDepth': ws['B42'].value,
            'columnHeight': ws['B43'].value,
        }
        for key, value in additional.items():
            print(f"{key}: {value}")

    # Get cost summary from SUMMARY - NEW tab
    if 'SUMMARY - NEW' in wb.sheetnames:
        ws = wb['SUMMARY - NEW']
        print("\n" + "=" * 100)
        print("### COST SUMMARY FROM SUMMARY - NEW TAB ###")
        print("=" * 100 + "\n")

        costs = {}
        # Extract all totals from column D
        cost_rows = [
            ('PLANS & ENGINEERING TOTAL', 5),
            ('LAYOUT TOTAL', 7),
            ('PERMIT TOTAL', 9),
            ('EXCAVATION TOTAL', 11),
            ('PLUMBING TOTAL', 13),
            ('GAS TOTAL', 15),
            ('STEEL TOTAL', 17),
            ('ELECTRICAL TOTAL', 19),
            ('SHOTCRETE LABOR TOTAL', 21),
            ('SHOTCRETE MATERIAL TOTAL', 23),
            ('TILE LABOR TOTAL', 25),
            ('TILE MATERIAL TOTAL', 27),
            ('COPING / DECKING LABOR TOTAL', 29),
            ('COPING / DECKING MATERIAL TOTAL', 31),
            ('STONE / ROCKWORK LABOR TOTAL', 33),
            ('STONE / ROCKWORK MATERIAL TOTAL', 35),
            ('EQUIPMENT ORDERED TOTAL', 39),
            ('EQUIPMENT SET TOTAL', 41),
            ('CLEAN-UP TOTAL', 43),
            ('INTERIOR FINISH TOTAL', 45),
            ('WATER TRUCK TOTAL', 47),
            ('FIBERGLASS SHELL TOTAL', 49),
            ('FIBERGLASS INSTALL TOTAL', 51),
            ('START-UP / ORIENTATION TOTAL', 53),
            ('TOTAL COST', 71),
        ]

        for label, row in cost_rows:
            value = ws[f'D{row}'].value
            costs[label] = value if value is not None else 0
            print(f"{label}: ${value if value is not None else 0:,.2f}")

    # Get retail price from NEW POOL tab
    if 'NEW POOL' in wb.sheetnames:
        ws = wb['NEW POOL']
        print("\n### PRICING FROM NEW POOL TAB ###\n")
        retail_price = ws['D181'].value
        sale_price = ws['B191'].value
        print(f"RETAIL PRICE: ${retail_price if retail_price is not None else 0:,.2f}")
        print(f"SALE PRICE: ${sale_price if sale_price is not None else 0:,.2f}")

    wb.close()

    return inputs, costs

if __name__ == "__main__":
    # Extract from Ashley Drennen file
    inputs, costs = extract_inputs_and_costs(r"C:\dev\PPAS\Ashley Drennen.xlsx")

    print("\n" + "=" * 100)
    print("### ANALYSIS COMPLETE ###")
    print("=" * 100)
