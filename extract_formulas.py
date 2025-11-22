import openpyxl
import sys

# Open the Master Excel Pricing workbook WITHOUT data_only to see formulas
wb = openpyxl.load_workbook('Master Excel Pricing.xlsx', data_only=False)

# Find the COST-NEW sheet
cost_new_sheet = None
for sheet_name in wb.sheetnames:
    if 'cost' in sheet_name.lower() and 'new' in sheet_name.lower():
        cost_new_sheet = wb[sheet_name]
        break

if not cost_new_sheet:
    print("COST-NEW sheet not found! Available sheets:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")
    sys.exit(1)

print(f"Reading formulas from sheet: {cost_new_sheet.title}")
print("=" * 100)

# Key categories to extract
categories = {
    'PLUMBING': {'start': None, 'formulas': {}},
    'ELECTRICAL': {'start': None, 'formulas': {}},
    'SHOTCRETE': {'start': None, 'formulas': {}},
    'EQUIPMENT SET': {'start': None, 'formulas': {}},
    'EXCAVATION': {'start': None, 'formulas': {}},
}

# Scan for category headers and their formulas
for row_idx in range(1, 500):
    cell_a = cost_new_sheet[f'A{row_idx}']
    if cell_a.value and isinstance(cell_a.value, str):
        upper_val = cell_a.value.upper()

        # Check if this is a category we're tracking
        for cat_name in categories.keys():
            if cat_name in upper_val and categories[cat_name]['start'] is None:
                categories[cat_name]['start'] = row_idx
                print(f"\n{'='*100}")
                print(f"CATEGORY: {cat_name} (starting at row {row_idx})")
                print(f"{'='*100}")

                # Read next 40 rows to capture formulas
                for offset in range(1, 41):
                    check_row = row_idx + offset
                    desc_cell = cost_new_sheet[f'A{check_row}']

                    # Look for cells with formulas in columns B-F
                    if desc_cell.value:
                        desc = str(desc_cell.value)

                        # Skip if it's another category or total
                        if any(c in desc.upper() for c in ['CATEGORY:', 'TOTAL']) and 'TOTAL' not in desc.upper():
                            break

                        formulas_found = {}
                        for col in ['B', 'C', 'D', 'E', 'F']:
                            cell = cost_new_sheet[f'{col}{check_row}']
                            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                                formulas_found[col] = cell.value

                        if formulas_found:
                            print(f"\nRow {check_row} - {desc[:50]}")
                            for col, formula in formulas_found.items():
                                print(f"  [{col}]: {formula}")

print("\n" + "="*100)
print("FORMULA EXTRACTION COMPLETE")
print("="*100)
