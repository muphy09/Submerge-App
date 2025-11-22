import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import json
import re

def analyze_excel(file_path):
    """Comprehensively analyze the Excel workbook"""
    wb = openpyxl.load_workbook(file_path, data_only=False)

    print("=" * 80)
    print(f"ANALYZING: {file_path}")
    print("=" * 80)

    # 1. List all sheets
    print("\n### SHEETS IN WORKBOOK ###")
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"{idx}. {sheet_name}")

    # 2. Analyze NEW POOL tab
    if 'NEW POOL' in wb.sheetnames:
        print("\n" + "=" * 80)
        print("### NEW POOL TAB ANALYSIS ###")
        print("=" * 80)
        ws = wb['NEW POOL']

        print("\n--- Input Fields (cells with values and nearby labels) ---")
        input_fields = []
        for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=20):
            for cell in row:
                # Look for cells that might be labels (text) followed by input cells
                if cell.value and isinstance(cell.value, str) and cell.value.strip():
                    label = cell.value.strip()
                    # Check adjacent cells for values or formulas
                    next_col_letter = get_column_letter(cell.column + 1)
                    next_cell = ws[f'{next_col_letter}{cell.row}']

                    if next_cell.value is not None or next_cell.data_type == 'f':
                        input_fields.append({
                            'label': label,
                            'label_cell': cell.coordinate,
                            'value_cell': next_cell.coordinate,
                            'value': next_cell.value,
                            'is_formula': next_cell.data_type == 'f'
                        })

        for field in input_fields[:50]:  # Show first 50
            formula_indicator = " [FORMULA]" if field['is_formula'] else ""
            print(f"  {field['label']}: {field['value_cell']} = {field['value']}{formula_indicator}")

    # 3. Analyze COST-NEW tab
    if 'COST-NEW' in wb.sheetnames:
        print("\n" + "=" * 80)
        print("### COST-NEW TAB ANALYSIS ###")
        print("=" * 80)
        ws = wb['COST-NEW']

        print("\n--- All formulas and calculations ---")
        formulas = []
        for row in ws.iter_rows(min_row=1, max_row=200):
            for cell in row:
                if cell.data_type == 'f':
                    # Get the label from the cell to the left if it exists
                    label_col_letter = get_column_letter(cell.column - 1) if cell.column > 1 else None
                    label = ""
                    if label_col_letter:
                        label_cell = ws[f'{label_col_letter}{cell.row}']
                        if label_cell.value:
                            label = str(label_cell.value).strip()

                    formulas.append({
                        'cell': cell.coordinate,
                        'label': label,
                        'formula': cell.value
                    })

        for formula in formulas[:100]:  # Show first 100
            print(f"  {formula['cell']}: {formula['label']}")
            print(f"    Formula: {formula['formula']}")
            print()

    # 4. Find cells with specific keywords related to cost
    print("\n" + "=" * 80)
    print("### SEARCHING FOR KEY COST TERMS ###")
    print("=" * 80)

    keywords = ['total', 'subtotal', 'cost', 'price', 'labor', 'material', 'job cost', 'grand total']

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- {sheet_name} ---")
        found_any = False

        for row in ws.iter_rows(min_row=1, max_row=200):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_lower = cell.value.lower()
                    for keyword in keywords:
                        if keyword in cell_lower:
                            found_any = True
                            # Get the value cell (usually to the right)
                            value_col_letter = get_column_letter(cell.column + 1)
                            value_cell = ws[f'{value_col_letter}{cell.row}']

                            formula_str = ""
                            if value_cell.data_type == 'f':
                                formula_str = f" [Formula: {value_cell.value}]"

                            print(f"  {cell.coordinate}: '{cell.value}' -> {value_cell.coordinate} = {value_cell.value}{formula_str}")
                            break

        if not found_any:
            print("  (No key cost terms found)")

    # 5. Analyze all other sheets for reference data
    print("\n" + "=" * 80)
    print("### OTHER SHEETS STRUCTURE ###")
    print("=" * 80)

    for sheet_name in wb.sheetnames:
        if sheet_name not in ['NEW POOL', 'COST-NEW']:
            print(f"\n--- {sheet_name} ---")
            ws = wb[sheet_name]

            # Get headers (first few rows)
            print("  Headers/First rows:")
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, max_col=15), 1):
                row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
                if any(row_values):
                    print(f"    Row {row_idx}: {' | '.join(row_values[:10])}")

    wb.close()

if __name__ == "__main__":
    analyze_excel(r"C:\dev\PPAS\Master Excel Pricing.xlsx")
