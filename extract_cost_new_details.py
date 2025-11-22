import openpyxl
import sys

# Open the Ashley Drennen workbook
wb = openpyxl.load_workbook('Ashley Drennen.xlsx', data_only=True)

# Find the COST-NEW sheet
cost_new_sheet = None
for sheet_name in wb.sheetnames:
    if 'cost' in sheet_name.lower() and 'new' in sheet_name.lower():
        cost_new_sheet = wb[sheet_name]
        break

if not cost_new_sheet:
    print("COST-NEW sheet not found!")
    sys.exit(1)

print(f"Reading from sheet: {cost_new_sheet.title}")
print("=" * 80)

# Read all rows and look for categories and their line items
current_category = None
for row_idx, row in enumerate(cost_new_sheet.iter_rows(min_row=1, max_row=500, values_only=True), start=1):
    # Check if this row has data
    if not any(row):
        continue

    # Get first few cells
    col_a = row[0] if len(row) > 0 else None
    col_b = row[1] if len(row) > 1 else None
    col_c = row[2] if len(row) > 2 else None
    col_d = row[3] if len(row) > 3 else None
    col_e = row[4] if len(row) > 4 else None
    col_f = row[5] if len(row) > 5 else None

    # Check if this is a category header (bold text or specific keywords)
    if col_a and isinstance(col_a, str):
        upper_a = col_a.upper()
        # Check for category headers
        if any(cat in upper_a for cat in ['EXCAVATION', 'PLUMBING', 'STEEL', 'ELECTRICAL',
                                           'SHOTCRETE', 'TILE', 'COPING', 'DECKING', 'EQUIPMENT',
                                           'INTERIOR', 'CLEANUP', 'STONE', 'ROCKWORK', 'PLANS',
                                           'ENGINEERING', 'LAYOUT', 'PERMIT', 'GAS', 'DRAINAGE',
                                           'WATER FEATURES', 'STARTUP', 'ORIENTATION']):
            if 'TOTAL' not in upper_a:  # Don't treat totals as categories
                current_category = col_a
                print(f"\n{'='*80}")
                print(f"CATEGORY: {current_category}")
                print(f"{'='*80}")
                print(f"{'Description':<40} {'Quantity':>12} {'Unit Price':>15} {'Total':>15}")
                print("-" * 80)

    # Print line items (rows with description and numbers)
    if current_category and col_a and isinstance(col_a, str):
        # Try to find quantity, unit price, and total in the row
        description = col_a

        # Skip if it's a category header or total row
        if any(word in description.upper() for word in ['TOTAL', 'SUBTOTAL', 'CATEGORY']):
            if 'TOTAL' in description.upper():
                print(f"{'-'*80}")
                # Find the total value in the row
                for val in row[1:]:
                    if isinstance(val, (int, float)) and val != 0:
                        print(f"{'TOTAL':<40} {'':<12} {'':<15} ${val:>14,.2f}")
                        break
            continue

        # Look for numeric values in the row
        values = [v for v in row[1:] if isinstance(v, (int, float)) and v != 0]

        if len(values) >= 1:
            # Assume format: Description | Quantity | Unit Price | Total (or variations)
            if len(values) == 1:
                # Only total
                print(f"{description:<40} {'':<12} {'':<15} ${values[0]:>14,.2f}")
            elif len(values) == 2:
                # Quantity and Total, or Unit Price and Total
                print(f"{description:<40} {values[0]:>12.2f} {'':<15} ${values[1]:>14,.2f}")
            elif len(values) >= 3:
                # Quantity, Unit Price, and Total
                qty = values[0]
                unit_price = values[1]
                total = values[2]
                print(f"{description:<40} {qty:>12.2f} ${unit_price:>14.2f} ${total:>14,.2f}")

print("\n" + "="*80)
print("EXTRACTION COMPLETE")
print("="*80)
