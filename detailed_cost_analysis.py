import openpyxl
from openpyxl.utils import get_column_letter
import json

def detailed_cost_analysis(file_path):
    """Extract detailed cost calculations from COST-NEW tab"""
    wb = openpyxl.load_workbook(file_path, data_only=False)

    print("=" * 100)
    print("DETAILED COST-NEW TAB ANALYSIS")
    print("=" * 100)

    if 'COST - NEW' not in wb.sheetnames:
        print("COST - NEW tab not found!")
        return

    ws = wb['COST - NEW']

    # Extract all rows with formulas and their labels
    sections = []
    current_section = None

    for row_idx in range(1, 400):  # Scan first 400 rows
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, min_col=1, max_col=10))[0]
        cell_a = row[0]
        cell_b = row[1] if len(row) > 1 else None
        cell_c = row[2] if len(row) > 2 else None
        cell_d = row[3] if len(row) > 3 else None

        # Check if this is a section header (TOTAL in column A)
        if cell_a.value and isinstance(cell_a.value, str) and 'TOTAL' in cell_a.value.upper():
            if current_section:
                sections.append(current_section)

            current_section = {
                'name': cell_a.value,
                'row': row_idx,
                'items': []
            }
        elif current_section:
            # This is a detail row within the section
            label = str(cell_a.value) if cell_a.value else ""
            qty = cell_c.value if cell_c else None
            cost = cell_d.value if cell_d else None

            if label or qty or cost:
                current_section['items'].append({
                    'row': row_idx,
                    'label': label,
                    'price': cell_b.value if cell_b else None,
                    'qty': qty,
                    'qty_is_formula': cell_c.data_type == 'f' if cell_c else False,
                    'qty_formula': qty if (cell_c and cell_c.data_type == 'f') else None,
                    'cost': cost,
                    'cost_is_formula': cell_d.data_type == 'f' if cell_d else False,
                    'cost_formula': cost if (cell_d and cell_d.data_type == 'f') else False,
                })

    if current_section:
        sections.append(current_section)

    # Print all sections
    for section in sections:
        print(f"\n{'='*100}")
        print(f"SECTION: {section['name']} (Row {section['row']})")
        print(f"{'='*100}")

        for item in section['items']:
            print(f"\nRow {item['row']}: {item['label']}")
            if item['price']:
                print(f"  Price: {item['price']}")
            if item['qty_is_formula']:
                print(f"  Qty: {item['qty_formula']} [FORMULA]")
            elif item['qty']:
                print(f"  Qty: {item['qty']}")
            if item['cost_is_formula']:
                print(f"  Cost: {item['cost_formula']} [FORMULA]")
            elif item['cost']:
                print(f"  Cost: {item['cost']}")

    # Now let's get the actual TOTAL COST cell
    print("\n" + "=" * 100)
    print("FINDING TOTAL COST CELL")
    print("=" * 100)

    for row in ws.iter_rows(min_row=1, max_row=400):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if 'TOTAL COST' in cell.value.upper() or 'GRAND TOTAL' in cell.value.upper():
                    # Get the adjacent cells
                    next_col = get_column_letter(cell.column + 1)
                    next_next_col = get_column_letter(cell.column + 2)
                    next_next_next_col = get_column_letter(cell.column + 3)

                    value_cell = ws[f'{next_col}{cell.row}']
                    value_cell2 = ws[f'{next_next_col}{cell.row}']
                    value_cell3 = ws[f'{next_next_next_col}{cell.row}']

                    print(f"\n{cell.coordinate}: '{cell.value}'")
                    print(f"  {value_cell.coordinate}: {value_cell.value}")
                    print(f"  {value_cell2.coordinate}: {value_cell2.value}")
                    print(f"  {value_cell3.coordinate}: {value_cell3.value}")

    # Get summary cells from SUMMARY - NEW
    print("\n" + "=" * 100)
    print("SUMMARY - NEW TAB")
    print("=" * 100)

    if 'SUMMARY - NEW' in wb.sheetnames:
        ws_summary = wb['SUMMARY - NEW']

        for row in ws_summary.iter_rows(min_row=1, max_row=100):
            cell_a = row[0]
            if cell_a.value and isinstance(cell_a.value, str):
                if 'TOTAL' in cell_a.value.upper():
                    # Get cells in columns D and E
                    cell_d = row[3] if len(row) > 3 else None
                    cell_e = row[4] if len(row) > 4 else None

                    print(f"\n{cell_a.coordinate}: {cell_a.value}")
                    if cell_d:
                        print(f"  {cell_d.coordinate}: {cell_d.value}")
                    if cell_e:
                        print(f"  {cell_e.coordinate}: {cell_e.value}")

    wb.close()

if __name__ == "__main__":
    detailed_cost_analysis(r"C:\dev\PPAS\Master Excel Pricing.xlsx")
